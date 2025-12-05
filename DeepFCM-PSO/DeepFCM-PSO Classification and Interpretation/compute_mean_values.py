import pandas as pd
import numpy as np
from openpyxl import load_workbook

def compute_mean_deviations(num_dimensions, fold, column_names):
    input_file_names = [f"data{i}.xlsx" for i in range(1, fold+1)]

    # Initialize a sum DataFrame with zeros
    sum_df = pd.DataFrame(0, index=range(num_dimensions), columns=range(num_dimensions), dtype=float)

    # Read data from each input file and accumulate the sum
    dfs = []
    for file_name in input_file_names:
        try:
            df = pd.read_excel(file_name, header=None)
            dfs.append(df.iloc[1:, :].astype(float))
            sum_df += dfs[-1]
        except Exception as e:
            print(f"Error occurred while processing {file_name}: {e}")

    # Calculate the mean DataFrame
    num_files = len(input_file_names)
    mean_df = sum_df / num_files
    mean_df.to_excel("mean_values.xlsx", index=False, header=False)
    print("Mean values saved to mean_values.xlsx")

    # Calculate standard deviations
    deviations = np.std([df.values for df in dfs], axis=0)
    deviation_df = pd.DataFrame(deviations)
    deviation_df.to_excel('deviations.xlsx', index=False, header=False)
    print("Deviations values saved to deviations.xlsx")

    # Load the destination Excel files
    destination_wb_mean = load_workbook('mean_values.xlsx')
    destination_wb_deviation = load_workbook('deviations.xlsx')

    # Access the first sheet of each workbook
    destination_sheet_mean = destination_wb_mean.active
    destination_sheet_deviation = destination_wb_deviation.active

    # Insert column names as the first row in mean and deviations sheets
    for index, value in enumerate(column_names, start=1):
        destination_sheet_mean.cell(row=1, column=index, value=value)
        destination_sheet_deviation.cell(row=1, column=index, value=value)

    # Save the changes to the destination files
    destination_wb_mean.save('mean_values.xlsx')
    destination_wb_deviation.save('deviations.xlsx')

